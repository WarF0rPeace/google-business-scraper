import openpyxl
from openpyxl.styles import Font, PatternFill
from modules.exporter import Exporter
from modules.logger import get_logger

logger = get_logger(__name__)

class ExcelExporter(Exporter):
    def export(self, data, file_path):
        if not data:
            logger.warning("No data provided for export. Excel file was not created.")
            return

        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active

            headers = data[0].keys()
            sheet.append(list(headers))

            header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            header_font = Font(bold=True)

            for col_num, header in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font

            for row_data in data:
                sheet.append(list(row_data.values()))

            for col_num, header in enumerate(headers, 1):
                column_letter = openpyxl.utils.get_column_letter(col_num)
                max_length = len(header)
                for row in data:
                    cell_value = str(row[header])
                    max_length = max(max_length, len(cell_value))

                sheet.column_dimensions[column_letter].width = max_length + 2

            sheet.auto_filter.ref = sheet.dimensions
            workbook.save(file_path)
            logger.info(f"Data successfully exported to Excel at {file_path}.")
        except Exception as e:
            logger.error(f"Failed to export data to Excel at {file_path}: {e}", exc_info=True)
